import streamlit as st
import pandas as pd
import re
import math
import time
import zipfile
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook


# =========================
# Settings
# =========================
DETAIL_LIMIT = 9                # detail_1 ~ detail_9
DEFAULT_ID_COL = "A"            # A파일 상품아이디 기본 열
DEFAULT_CHUNK = 100             # 100행 단위 분할
DEFAULT_OUT_SHEET_INDEX = 0     # b.xlsx 템플릿에서 값을 쓸 시트 (0=첫 시트)
PROPOSAL_FORMAT_LABEL = "라스트오더 상품제안서"

# =========================
# Utils
# =========================
def col_idx(col: str) -> int:
    """Excel column letters -> 0-based index (supports AA, AB, AC...)"""
    idx = 0
    for c in col.upper():
        idx = idx * 26 + (ord(c) - ord("A") + 1)
    return idx - 1


def extract_bracket_items(val):
    """
    - [ ... ] 블록이 여러 개면 각각 아이템으로 추출
    - 블록 내부에 콤마가 있으면 추가 분리
    - 대괄호가 없으면 콤마 분리로 폴백
    """
    if pd.isna(val):
        return []
    s = str(val).strip()
    if not s:
        return []

    blocks = re.findall(r"\[([^\]]+)\]", s)
    items = []

    if blocks:
        for blk in blocks:
            blk = blk.strip()
            if not blk:
                continue
            if "," in blk:
                items.extend([p.strip() for p in blk.split(",") if p.strip()])
            else:
                items.append(blk)
    else:
        s2 = s.replace("[", "").replace("]", "").strip()
        if not s2:
            return []
        items = [p.strip() for p in s2.split(",") if p.strip()]

    return items


def uniq_keep_order(seq):
    return list(dict.fromkeys(seq))


def normalize_header(val):
    if pd.isna(val):
        return ""
    return re.sub(r"\s+", "", str(val)).strip().lower()


def find_header_col(header_values, *aliases):
    normalized_aliases = {normalize_header(alias) for alias in aliases}
    for idx, header in enumerate(header_values):
        if normalize_header(header) in normalized_aliases:
            return idx
    return None


def split_option_values(val):
    if pd.isna(val):
        return [""]

    text = str(val).strip()
    if not text or text.lower() == "nan" or text in {"단품", "상동"}:
        return [""]

    options = [part.strip() for part in re.split(r"[,，\n]+", text) if part.strip()]
    return options or [""]


def resolve_same_as_above(val, previous):
    if pd.isna(val):
        return val

    text = str(val).strip()
    if text == "상동":
        return previous
    return val


def normalize_proposal_df(raw: pd.DataFrame):
    """
    라스트오더 상품제안서 양식:
    - 1행은 안내 제목, 2행이 실제 헤더
    - 기존 변환기는 고정 열 위치(A/C/D/E/H/I/J/M/P/S/T)를 참조하므로
      상품제안서를 그 내부 A형식으로 변환해 기존 로직을 재사용한다.
    """
    required_headers = {"브랜드", "상품명", "재고수량", "소비자가", "판매가", "공급가"}
    header_row_idx = None

    for idx in range(min(len(raw), 10)):
        normalized = {normalize_header(v) for v in raw.iloc[idx].tolist()}
        if {normalize_header(v) for v in required_headers}.issubset(normalized):
            header_row_idx = idx
            break

    if header_row_idx is None:
        return None

    header_values = raw.iloc[header_row_idx].tolist()
    col_partner = find_header_col(header_values, "파트너사명")
    col_brand = find_header_col(header_values, "브랜드")
    col_name = find_header_col(header_values, "상품명")
    col_option = find_header_col(header_values, "구성(옵션)", "구성")
    col_expiry = find_header_col(header_values, "유통기한")
    col_stock = find_header_col(header_values, "재고수량")
    col_consumer_price = find_header_col(header_values, "소비자가")
    col_sale_price = find_header_col(header_values, "판매가")
    col_supply_price = find_header_col(header_values, "공급가")
    col_tax = find_header_col(header_values, "면/과세", "면과세")
    col_url = find_header_col(header_values, "판매링크URL", "판매링크 URL")

    rows = raw.iloc[header_row_idx + 1:].reset_index(drop=True)
    normalized_rows = []
    extra_rows = []
    previous_tax = ""

    for _, source_row in rows.iterrows():
        product_name = source_row.iloc[col_name] if col_name is not None else None
        if pd.isna(product_name) or not str(product_name).strip():
            continue

        partner = source_row.iloc[col_partner] if col_partner is not None else ""
        brand = source_row.iloc[col_brand] if col_brand is not None else ""
        option_value = source_row.iloc[col_option] if col_option is not None else ""
        expiry = source_row.iloc[col_expiry] if col_expiry is not None else ""
        stock = source_row.iloc[col_stock] if col_stock is not None else ""
        consumer_price = source_row.iloc[col_consumer_price] if col_consumer_price is not None else ""
        sale_price = source_row.iloc[col_sale_price] if col_sale_price is not None else ""
        supply_price = source_row.iloc[col_supply_price] if col_supply_price is not None else ""
        tax = source_row.iloc[col_tax] if col_tax is not None else ""
        sale_url = source_row.iloc[col_url] if col_url is not None else ""
        tax = resolve_same_as_above(tax, previous_tax)
        if pd.notna(tax) and str(tax).strip():
            previous_tax = tax

        pid = "^|^".join(
            str(v).strip()
            for v in [partner, brand, product_name]
            if pd.notna(v) and str(v).strip()
        )

        for option in split_option_values(option_value):
            row = ["" for _ in range(col_idx("T") + 1)]
            row[col_idx("A")] = pid
            row[col_idx("D")] = product_name
            row[col_idx("E")] = option
            row[col_idx("H")] = consumer_price
            row[col_idx("I")] = supply_price
            row[col_idx("J")] = sale_price
            row[col_idx("M")] = stock
            row[col_idx("P")] = expiry
            normalized_rows.append(row)
            extra_rows.append({
                "_proposal_brand": brand,
                "_proposal_tax": tax,
                "_proposal_sale_url": sale_url,
            })

    if not normalized_rows:
        return None

    normalized_df = pd.DataFrame(normalized_rows)
    for extra_key in ["_proposal_brand", "_proposal_tax", "_proposal_sale_url"]:
        normalized_df[extra_key] = [extra[extra_key] for extra in extra_rows]

    normalized_df.attrs["source_format"] = "proposal"
    normalized_df.attrs["source_format_label"] = PROPOSAL_FORMAT_LABEL
    return normalized_df


def read_a_workbook(file_bytes: bytes, requested_id_col_letter: str):
    raw_df = pd.read_excel(BytesIO(file_bytes), header=None)
    proposal_df = normalize_proposal_df(raw_df)
    if proposal_df is not None:
        return proposal_df, "A"

    default_df = pd.read_excel(BytesIO(file_bytes))
    default_df.attrs["source_format"] = "default"
    default_df.attrs["source_format_label"] = "기본 A파일"
    return default_df, requested_id_col_letter


def build_bd_from_group(a: pd.DataFrame, pid_series: pd.Series, pid_value: str, one_line: bool):
    """
    BD 규칙:
    - main: S에서 첫 유효 main 1개
    - detail: T에서 전체를 모아 중복 제거 후 detail_1~detail_9
    - one_line=True면 줄바꿈 없이 한 줄로 합침(옵션 상품에서 사용)
    """
    mask = (pid_series == pid_value).to_numpy()

    s_all = a.iloc[:, col_idx("S")]
    t_all = a.iloc[:, col_idx("T")]

    main_candidates = []
    for sv in s_all[mask]:
        main_candidates.extend(extract_bracket_items(sv))
    main = main_candidates[0] if main_candidates else ""

    detail_items = []
    for tv in t_all[mask]:
        detail_items.extend(extract_bracket_items(tv))

    detail_items = uniq_keep_order([x for x in detail_items if x])[:DETAIL_LIMIT]

    parts = []
    if main:
        parts.append(f"main^|^https://m.lastorder.in/{main}")
    for i, it in enumerate(detail_items, start=1):
        parts.append(f"detail_{i}^|^https://m.lastorder.in/{it}")

    return " ".join(parts) if one_line else "\n".join(parts)


def join_unique(vals):
    cleaned = [str(v).strip() for v in vals if pd.notna(v) and str(v).strip()]
    return "^|^".join(uniq_keep_order(cleaned))


def validate_a_df(a: pd.DataFrame, id_col_letter: str):
    """
    현재 로직에서 참조하는 A파일 컬럼(엑셀 위치 기준)
    C, D, E, H, I, J, M, P, S, T, + 상품아이디(id_col_letter)
    """
    required = ["C", "D", "E", "H", "I", "J", "M", "P", "S", "T", id_col_letter]
    max_needed = max(col_idx(c) for c in required)
    if a.shape[1] <= max_needed:
        missing = [c for c in required if col_idx(c) >= a.shape[1]]
        return False, f"A파일 컬럼 부족: {missing} (현재 컬럼 수: {a.shape[1]})"
    return True, ""


def make_b_rows_from_a(a: pd.DataFrame, id_col_letter: str):
    """
    A -> B (행 리스트/행 데이터)
    - 상품아이디 중복이면 옵션그룹
    - 같은 상품아이디는 1행으로 묶음
    - 옵션그룹:
        AH='y', AI='선택'
        AJ = 옵션값(E) ^|^ 연결 (중복 제거, 등장 순)
        AN = 옵션재고(M) : AJ 옵션값 순서에 맞춰 매칭한 뒤 ^|^ 연결
        O/Q/AW(판매종료일) = 그룹 최소 날짜
        BD = 그룹 전체 이미지 합치기 + (옵션일 경우) 한 줄(one_line)
    - 비옵션:
        O/Q/AW는 그룹 최소(결과적으로 단일이라 동일)
        BD는 줄바꿈 유지
    - 매칭 뒤틀림 방지: 대표행/Series는 reset_index(drop=True) + to_numpy 사용
    """
    is_proposal_format = a.attrs.get("source_format") == "proposal"
    pid = a.iloc[:, col_idx(id_col_letter)].astype(str).fillna("").str.strip()
    dup_mask = pid.duplicated(keep=False)

    # 대표행(각 pid 첫 행) - 인덱스 리셋이 매우 중요(align 방지)
    rep_mask = ~pid.duplicated(keep="first")
    a_rep = a.loc[rep_mask].reset_index(drop=True)
    pid_rep = pid.loc[rep_mask].reset_index(drop=True)

    option_pids = set(pid[dup_mask])

    # 판매종료일 그룹 최소
    p_all_dt = pd.to_datetime(a.iloc[:, col_idx("P")], errors="coerce")
    p_min_map = p_all_dt.groupby(pid).min()

    # 옵션값(E)
    e_series = a.iloc[:, col_idx("E")]

    # 옵션값(AJ) 맵 (중복 제거, 순서 유지)
    opt_value_map = e_series.groupby(pid, sort=False).apply(join_unique).to_dict()

    # 옵션재고(AN): A파일 M열 기반, 옵션값 순서에 맞춰 매칭
    m_stock_series = a.iloc[:, col_idx("M")]
    opt_stock_map = {}
    df_opt = pd.DataFrame({"pid": pid, "opt": e_series, "stk": m_stock_series})

    for pid_val, grp in df_opt.groupby("pid", sort=False):
        # 옵션값 등장 순서 + 중복 제거
        opt_vals_raw = [str(v).strip() for v in grp["opt"].tolist() if pd.notna(v) and str(v).strip()]
        opt_vals = uniq_keep_order(opt_vals_raw)

        stocks_out = []
        for ov in opt_vals:
            sub = grp.loc[grp["opt"].astype(str).str.strip() == ov, "stk"]

            chosen = ""
            for sv in sub.tolist():
                if pd.isna(sv):
                    continue
                ss = str(sv).strip()
                if ss != "" and ss.lower() != "nan":
                    chosen = ss
                    break
            stocks_out.append(chosen)

        opt_stock_map[pid_val] = "^|^".join(stocks_out)

    # 결과 행 생성 (B 템플릿에 쓸 “열문자:값” dict 리스트)
    out_rows = []
    for i in range(len(a_rep)):
        pid_i = pid_rep.iloc[i]
        is_option = pid_i in option_pids

        row = {}

        # 고정
        row["A"] = 1
        row["B"] = 2
        row["C"] = 217089
        row["D"] = 1011307
        row["K"] = "n"

        # 기본 매핑(대표행 기준)
        row["H"]  = a_rep.iloc[:, col_idx("D")].to_numpy()[i]  # 상품명
        if is_proposal_format and "_proposal_brand" in a_rep:
            row["J"] = a_rep["_proposal_brand"].to_numpy()[i]
        row["U"]  = a_rep.iloc[:, col_idx("I")].to_numpy()[i]
        row["V"]  = a_rep.iloc[:, col_idx("H")].to_numpy()[i]
        row["AC"] = a_rep.iloc[:, col_idx("M")].to_numpy()[i]  # 기존 요구: AC <- M (대표행)
        if is_proposal_format:
            row["AY"] = ""
            row["AZ"] = ""
            row["AX"] = a_rep["_proposal_tax"].to_numpy()[i] if "_proposal_tax" in a_rep else ""
        else:
            row["AY"] = a_rep.iloc[:, col_idx("C")].to_numpy()[i]
            row["AZ"] = a_rep.iloc[:, col_idx("C")].to_numpy()[i]

        # W = (H - J) 계산 후 "-1"
        h_val = pd.to_numeric(a_rep.iloc[:, col_idx("H")], errors="coerce").fillna(0).to_numpy()[i]
        j_val = pd.to_numeric(a_rep.iloc[:, col_idx("J")], errors="coerce").fillna(0).to_numpy()[i]
        row["W"] = f"{int(h_val - j_val)}-1"

        # 판매종료일: 그룹 최소
        pmin = p_min_map.get(pid_i, pd.NaT)
        if pd.isna(pmin):
            row["O"] = 1
            row["Q"] = "2999-12-31"
            row["AW"] = "2999-12-31"
        else:
            d = pd.Timestamp(pmin).strftime("%Y-%m-%d")
            row["O"] = 2
            row["Q"] = d
            row["AW"] = d

        # 옵션 관련
        if is_option:
            row["AH"] = "y"
            row["AI"] = "선택"
            row["AJ"] = opt_value_map.get(pid_i, "")
            row["AN"] = opt_stock_map.get(pid_i, "")  # ✅ 옵션재고 (A파일 M열)
            # 옵션이미지: 한 줄로 합침
            row["BD"] = build_bd_from_group(a, pid, pid_i, one_line=True)
        else:
            # 비옵션: BD 줄바꿈 유지
            row["BD"] = build_bd_from_group(a, pid, pid_i, one_line=False)

        out_rows.append(row)

    return out_rows


def apply_rows_to_template(template_bytes: bytes, rows: list[dict], sheet_index: int, start_row: int = 2):
    """
    b.xlsx 템플릿(전체 탭 유지)에 rows를 첫 시트(기본) start_row부터 값으로 기입
    - 기존 템플릿 구조/다른 시트 유지
    """
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb.worksheets[sheet_index]

    for i, row in enumerate(rows):
        excel_row = start_row + i
        for col_letter, val in row.items():
            ws[f"{col_letter}{excel_row}"] = val

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def split_rows(rows: list[dict], chunk_size: int):
    if chunk_size <= 0:
        return [rows]
    return [rows[i:i+chunk_size] for i in range(0, len(rows), chunk_size)] or [[]]


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="A→B 변환기(옵션/템플릿 유지/분할)", layout="wide")
st.title("📦 A파일 → B템플릿(b.xlsx) 자동 변환기 (옵션/이미지/재고/분할/리포트)")

with st.expander("업로드 팁", expanded=True):
    st.write(
        "- 폴더처럼 쓰려면 **폴더 안 xlsx 여러 개를 한 번에 드래그&드롭** 하세요.\n"
        "- `b.xlsx` 템플릿은 앱에서 업로드하거나, 서버/로컬 실행 시 같은 폴더에 둬도 됩니다.\n"
        "- `라스트오더 상품제안서` 양식도 A파일 업로드에 넣으면 자동으로 컬럼을 맞춥니다.\n"
        "- 결과는 **b.xlsx 템플릿의 모든 탭을 유지**한 상태로, 첫 시트에 값만 채웁니다.\n"
        "- 옵션 그룹은 상품아이디 중복으로 판단하며, 옵션 이미지(BD)는 한 줄로 합칩니다."
    )

# Sidebar controls
st.sidebar.header("설정")
id_col_letter = st.sidebar.text_input("A파일 상품아이디 컬럼(엑셀 문자)", value=DEFAULT_ID_COL).strip().upper()
chunk_size = st.sidebar.number_input("분할 저장(행)", min_value=10, max_value=5000, value=DEFAULT_CHUNK, step=10)
sheet_index = st.sidebar.number_input("템플릿에 쓸 시트 인덱스(0=첫 시트)", min_value=0, max_value=30, value=DEFAULT_OUT_SHEET_INDEX, step=1)

# Template uploader (optional)
template_file = st.file_uploader("B 템플릿(b.xlsx) 업로드 (선택: 업로드 안 하면 기본 b.xlsx 사용)", type=["xlsx"])

# A files uploader
a_files = st.file_uploader("A파일 업로드 (여러 개 가능)", type=["xlsx"], accept_multiple_files=True)

run_btn = st.button("🚀 변환 시작", disabled=not a_files)

def get_template_bytes():
    # If user uploaded template, use it. Otherwise try to read local ./b.xlsx
    if template_file is not None:
        return template_file.getvalue()
    # fallback: local file (works when running locally/server with b.xlsx in same folder)
    local_path = Path("b.xlsx")
    if local_path.exists():
        return local_path.read_bytes()
    raise FileNotFoundError("b.xlsx 템플릿이 없습니다. b.xlsx를 업로드하거나 앱 폴더에 두세요.")

if run_btn:
    t0 = time.time()
    st.info("처리 중...")

    try:
        template_bytes = get_template_bytes()
    except Exception as e:
        st.error(f"템플릿 로드 실패: {e}")
        st.stop()

    summary_rows = []
    error_rows = []

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for uf in a_files:
            if uf.name.startswith("~$"):
                continue

            started = time.time()
            status = "OK"
            msg = ""
            input_rows = 0
            out_files = 0
            detected_format = ""

            try:
                file_bytes = uf.getvalue()
                a_df, effective_id_col_letter = read_a_workbook(file_bytes, id_col_letter)
                detected_format = a_df.attrs.get("source_format_label", "기본 A파일")
                input_rows = len(a_df)

                ok, vmsg = validate_a_df(a_df, effective_id_col_letter)
                if not ok:
                    raise ValueError(vmsg)

                rows = make_b_rows_from_a(a_df, effective_id_col_letter)

                # split by chunk_size
                chunks = split_rows(rows, int(chunk_size))
                for idx, chunk in enumerate(chunks, start=1):
                    out_xlsx = apply_rows_to_template(
                        template_bytes=template_bytes,
                        rows=chunk,
                        sheet_index=int(sheet_index),
                        start_row=2
                    )
                    out_name = f"{Path(uf.name).stem}_part{idx:03d}.xlsx"
                    zf.writestr(out_name, out_xlsx)
                    out_files += 1

            except Exception as e:
                status = "FAIL"
                msg = str(e)
                error_rows.append({"file": uf.name, "reason": msg})

            elapsed = round(time.time() - started, 3)
            summary_rows.append({
                "file": uf.name,
                "format": detected_format,
                "status": status,
                "input_rows": input_rows,
                "output_files": out_files,
                "seconds": elapsed,
                "message": msg
            })

        # summary_report.csv
        summary_df = pd.DataFrame(summary_rows)
        zf.writestr("summary_report.csv", summary_df.to_csv(index=False).encode("utf-8-sig"))

        # errors.csv
        if error_rows:
            errors_df = pd.DataFrame(error_rows)
            zf.writestr("errors.csv", errors_df.to_csv(index=False).encode("utf-8-sig"))

    zip_buf.seek(0)
    total_sec = round(time.time() - t0, 2)

    st.success(f"✅ 완료! 총 소요 {total_sec}s")
    st.subheader("📊 요약 리포트")
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)

    if error_rows:
        st.subheader("⚠️ 에러")
        st.dataframe(pd.DataFrame(error_rows), use_container_width=True)

    st.download_button(
        "📦 결과 ZIP 다운로드 (엑셀 + 리포트 포함)",
        data=zip_buf,
        file_name="B_result.zip",
        mime="application/zip"
    )
